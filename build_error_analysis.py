"""Build the error-analysis spreadsheet for the dense RAG run.

Rigorous-only policy:
  - Categories 1/2 (retrieval miss vs hit-but-wrong) are computed from the
    retrieval logs and are exhaustive over all errors.
  - Affirmation direction is computed from the confidence ordering no<maybe<yes.
  - "maybe-gold" is a transparent PROXY flag for likely label-ambiguity, not a
    verified judgment.
  - The 15-case hand labels from failures.md are included verbatim and checked
    against the automatic category; conflicts are surfaced, not hidden.
Nothing subjective is presented as fact.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("results/raw_outputs/dense_rag.json") as f:
    dense = json.load(f)

errors = [r for r in dense if r["predicted"] != r["gold_answer"]]

# ---- 15-case hand labels from failures.md (verbatim) -----------------------
SAMPLE = {
    "15483019": "1. Retrieval miss",
    "9582182": "2. Hit, wrong generation",
    "12970636": "2. Hit, wrong generation",
    "23719685": "2. Hit, wrong generation",
    "15112004": "2. Hit, wrong generation",
    "27643961": "2. Hit, wrong generation",
    "18926458": "2. Hit, wrong generation",
    "11977907": "2. Hit, wrong generation",
    "10430303": "3. Synthesis-requiring",
    "23571528": "3. Synthesis-requiring",
    "24139705": "3. Synthesis-requiring",
    "25636371": "4. Ambiguous gold",
    "25987398": "4. Ambiguous gold",
    "20101129": "4. Ambiguous gold",
    "16816043": "4. Ambiguous gold",
}

RANK = {"no": 0, "maybe": 1, "yes": 2}


def affirmation(gold, pred):
    if gold in RANK and pred in RANK:
        if RANK[pred] > RANK[gold]:
            return "over"
        if RANK[pred] < RANK[gold]:
            return "under"
    return "n/a"


rows = []
for r in errors:
    pid = str(r["pubid"])
    hit = r["retrieval_hit"]
    primary = "2. Hit, wrong generation" if hit else "1. Retrieval miss"
    aff = affirmation(r["gold_answer"], r["predicted"])
    proxy = "likely-ambiguous" if r["gold_answer"] == "maybe" else ""
    sample = SAMPLE.get(pid, "")
    # label_check: does the hand sample agree with the automatic category?
    if not sample:
        check = ""
    else:
        sample_is_miss = sample.startswith("1.")
        auto_is_miss = not hit
        if sample_is_miss != auto_is_miss and (sample_is_miss or auto_is_miss):
            # only flag the miss-vs-hit axis, since 3/4 are sub-types of hit
            check = "CONFLICT"
        else:
            check = "agree"
    note = ""
    if pid == "15483019":
        note = ("failures.md labels this 'Retrieval miss' but retrieval_hit=True; "
                "true category is hit-but-wrong generation.")
    rows.append({
        "pubid": pid,
        "question": r["question"],
        "gold": r["gold_answer"],
        "pred": r["predicted"],
        "retrieval_hit": hit,
        "primary_category": primary,
        "affirmation": aff,
        "maybe_gold_proxy": proxy,
        "sample_label_failuresmd": sample,
        "label_check": check,
        "notes": note,
    })

# Flag the true miss explicitly in notes
for row in rows:
    if not row["retrieval_hit"] and row["pubid"] != "15483019":
        row["notes"] = ("TRUE Category-1 retrieval miss (only one in 200); "
                        "not written up in failures.md.")

# ---- styling helpers -------------------------------------------------------
HEAD = Font(bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor="2F5496")
CONFLICT_FILL = PatternFill("solid", fgColor="F8CBAD")
MISS_FILL = PatternFill("solid", fgColor="FFE699")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

wb = Workbook()

# =========================== Sheet 1: All Errors ============================
ws = wb.active
ws.title = "All Errors"
cols = [
    ("pubid", 11),
    ("question", 60),
    ("gold", 7),
    ("pred", 7),
    ("retrieval_hit", 12),
    ("primary_category", 24),
    ("affirmation", 12),
    ("maybe_gold_proxy", 16),
    ("sample_label_failuresmd", 24),
    ("label_check", 12),
    ("notes", 50),
]
for i, (name, w) in enumerate(cols, 1):
    c = ws.cell(1, i, name)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.border = BORDER
    c.alignment = WRAP
    ws.column_dimensions[get_column_letter(i)].width = w

for ri, row in enumerate(rows, 2):
    vals = [row[k] for k, _ in cols]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        c.border = BORDER
        c.alignment = WRAP if cols[ci - 1][0] in ("question", "notes") else TOP
    if row["label_check"] == "CONFLICT":
        ws.cell(ri, 10).fill = CONFLICT_FILL
    if not row["retrieval_hit"]:
        ws.cell(ri, 5).fill = MISS_FILL
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

# =========================== Sheet 2: Summary ===============================
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 60


def hdr(ws, r, text):
    c = ws.cell(r, 1, text)
    c.font = Font(bold=True, size=12, color="2F5496")


def line(ws, r, label, val, note=""):
    ws.cell(r, 1, label).alignment = TOP
    vc = ws.cell(r, 2, val)
    vc.alignment = TOP
    vc.font = Font(bold=True)
    n = ws.cell(r, 3, note)
    n.alignment = WRAP


n_hit_wrong = sum(1 for x in errors if x["retrieval_hit"])
n_miss = sum(1 for x in errors if not x["retrieval_hit"])
over = sum(1 for x in errors if x["retrieval_hit"] and affirmation(x["gold_answer"], x["predicted"]) == "over")
under = sum(1 for x in errors if x["retrieval_hit"] and affirmation(x["gold_answer"], x["predicted"]) == "under")
maybe_gold_err = sum(1 for x in errors if x["gold_answer"] == "maybe")

r = 1
hdr(ws2, r, "Dense RAG — Error Analysis Summary"); r += 1
ws2.cell(r, 1, "PubMedQA pqa_labeled, 200 questions, Llama 3.1 8B via Groq"); r += 2

hdr(ws2, r, "Counts (automatic, from retrieval logs)"); r += 1
line(ws2, r, "Total questions", 200); r += 1
line(ws2, r, "Correct", 137); r += 1
line(ws2, r, "Errors", 63, "exhaustive: 62 + 1"); r += 1
line(ws2, r, "  1. Retrieval miss", n_miss, "retrieval_hit = False"); r += 1
line(ws2, r, "  2. Hit, wrong generation", n_hit_wrong, "retrieval_hit = True"); r += 2

hdr(ws2, r, "Affirmation direction (within the 62 hit-but-wrong)"); r += 1
line(ws2, r, "  Over-affirmation", over, f"{over/n_hit_wrong*100:.1f}% — pred more affirmative than gold (no<maybe<yes)"); r += 1
line(ws2, r, "  Under-affirmation", under, f"{under/n_hit_wrong*100:.1f}%"); r += 2

hdr(ws2, r, "Proxy / sample (NOT rigorous counts)"); r += 1
line(ws2, r, "  maybe-gold errors (ambiguity proxy)", maybe_gold_err,
     "transparent proxy for likely label-noise; not a verified 'ambiguous' count"); r += 1
line(ws2, r, "  hand-labeled sample (failures.md)", 15,
     "Categories 3/4 exist ONLY for this 15-case sample, not all 63"); r += 1
line(ws2, r, "  sample/auto conflicts", 1,
     "PubID 15483019 labeled 'miss' in failures.md but retrieval_hit=True"); r += 2

hdr(ws2, r, "Data-integrity notes"); r += 1
ws2.cell(r, 1, "1. failures.md case study for 'Retrieval miss' (15483019) is mislabeled; "
               "that row is a generation error.").alignment = WRAP
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
ws2.cell(r, 1, "2. The real (and only) retrieval miss is PubID 11570976 "
               "('Is it Crohn's disease?', gold=maybe, pred=no), not written up.").alignment = WRAP
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
ws2.cell(r, 1, "3. Synthesis-required vs ambiguous-gold are content judgments; no rigorous "
               "all-63 count exists. Do a full manual pass if a complete 4-way table is needed.").alignment = WRAP
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1

# =========================== Sheet 3: Affirmation ===========================
ws3 = wb.create_sheet("Affirmation detail")
for i, (name, w) in enumerate([("gold", 10), ("predicted", 12), ("count", 8),
                               ("direction", 12)], 1):
    c = ws3.cell(1, i, name)
    c.font = HEAD; c.fill = HEAD_FILL; c.border = BORDER
    ws3.column_dimensions[get_column_letter(i)].width = w

from collections import Counter
pairs = Counter((x["gold_answer"], x["predicted"]) for x in errors if x["retrieval_hit"])
ri = 2
for (g, p), cnt in sorted(pairs.items(), key=lambda z: -z[1]):
    d = affirmation(g, p)
    for ci, v in enumerate([g, p, cnt, d], 1):
        cell = ws3.cell(ri, ci, v); cell.border = BORDER
        if d == "over":
            cell.fill = PatternFill("solid", fgColor="FCE4D6")
    ri += 1
ws3.freeze_panes = "A2"

wb.save("results/error_analysis.xlsx")
print("Wrote results/error_analysis.xlsx")
print(f"Rows: {len(rows)} errors | hit-wrong {n_hit_wrong} | miss {n_miss} | over {over} ({over/n_hit_wrong*100:.1f}%)")
